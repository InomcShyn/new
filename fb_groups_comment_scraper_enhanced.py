# fb_groups_comment_scraper_enhanced.py - Enhanced Facebook Groups Scraper
import time, random, threading, re, requests, pandas as pd
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.common.exceptions import WebDriverException, NoSuchElementException
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import json
import os

# ----------------------------
# Enhanced Helper Utils
# ----------------------------
def parse_cookies_to_list(cookie_str):
    cookies_list = []
    for pair in cookie_str.split(';'):
        pair = pair.strip()
        if '=' in pair:
            name, value = pair.split('=', 1)
            cookies_list.append({'name': name.strip(), 'value': value.strip(), 'domain': '.facebook.com'})
    return cookies_list

def parse_cookies_to_dict(cookie_str):
    d = {}
    for pair in cookie_str.split(';'):
        pair = pair.strip()
        if '=' in pair:
            name, value = pair.split('=', 1)
            d[name.strip()] = value.strip()
    return d

def clean_text(text):
    if not text:
        return ""
    text = re.sub(r'\s+', ' ', text.strip())
    # Remove UI elements but preserve Vietnamese text
    ui_patterns = [
        r'\b(Like|Reply|Share|Comment|Translate|Hide|Report|Block)\b',
        r'\b(Thích|Trả lời|Chia sẻ|Bình luận|Dịch|Ẩn|Báo cáo|Chặn)\b',
        r'\b\d+\s*(min|minutes?|hours?|days?|seconds?|phút|giờ|ngày|giây)\s*(ago|trước)?\b',
        r'\b(Top fan|Most relevant|Newest|All comments|Bình luận hàng đầu)\b'
    ]
    for pattern in ui_patterns:
        text = re.sub(pattern, '', text, flags=re.IGNORECASE)
    return text.strip()

def extract_user_info_from_url(url):
    """Enhanced user info extraction from Facebook URLs"""
    patterns = {
        'profile_id': r'profile\.php\?id=(\d+)',
        'username': r'facebook\.com/([^/?]+)',
        'user_id': r'/(\d{10,})/?$'
    }
    
    user_info = {}
    for key, pattern in patterns.items():
        match = re.search(pattern, url)
        if match:
            user_info[key] = match.group(1)
    
    return user_info

def validate_username(username):
    """Enhanced username validation"""
    if not username or len(username) < 2:
        return False
    
    # Remove common prefixes/suffixes
    username = re.sub(r'^(Mr\.|Ms\.|Mrs\.|Dr\.|Prof\.)\s*', '', username)
    username = re.sub(r'\s+(Jr\.|Sr\.|I|II|III|IV|V)$', '', username)
    
    # Check for valid characters
    if not re.match(r'^[a-zA-ZÀ-ỹ\s\-\.\']+$', username):
        return False
    
    # Check length
    if len(username) > 50:
        return False
    
    return True

def extract_comment_content(text, username=""):
    """Enhanced comment content extraction"""
    if not text:
        return ""
    
    # Remove username from beginning
    if username and text.startswith(username):
        text = text[len(username):].strip()
    
    # Remove username with word boundaries
    if username:
        text = re.sub(rf'\b{re.escape(username)}\b', '', text, count=1).strip()
    
    # Remove common prefixes
    prefixes_to_remove = [
        r'^[A-Z][a-z]+:\s*',  # "Name: "
        r'^@[a-zA-Z0-9_]+\s*',  # "@username "
        r'^Reply to [A-Z][a-z]+:\s*',  # "Reply to Name: "
    ]
    
    for prefix in prefixes_to_remove:
        text = re.sub(prefix, '', text)
    
    # Clean up extra whitespace
    text = re.sub(r'\s+', ' ', text).strip()
    
    return text

# ----------------------------
# Enhanced Facebook Groups Scraper
# ----------------------------
class EnhancedFacebookGroupsScraper:
    def __init__(self, cookie_str, headless=True):
        options = Options()
        if headless:
            options.add_argument("--headless=new")
        options.add_argument("--disable-notifications")
        options.add_argument("--disable-popup-blocking")
        options.add_argument("--disable-gpu")
        options.add_argument("--no-sandbox")
        options.add_argument("--disable-dev-shm-usage")
        options.add_argument("--disable-blink-features=AutomationControlled")
        
        # User agent that works well with Facebook groups
        options.add_argument("user-agent=Mozilla/5.0 (Android 10; Mobile; rv:109.0) Gecko/111.0 Firefox/109.0")
        options.add_argument("window-size=414,896")
        options.add_experimental_option("excludeSwitches", ["enable-automation", "enable-logging"])
        options.add_experimental_option('useAutomationExtension', False)
        
        self.driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)
        self.wait = WebDriverWait(self.driver, 15)
        self.cookie_str = cookie_str or ""
        self.cookies_list = parse_cookies_to_list(self.cookie_str)
        self.cookies_dict = parse_cookies_to_dict(self.cookie_str)
        self._stop_flag = False
        self.current_layout = None
        self.user_cache = {}  # Cache for user information
        
        if self.cookies_list:
            self._login_with_cookies()

    def _login_with_cookies(self):
        # Start with mobile Facebook for groups
        self.driver.get("https://m.facebook.com")
        time.sleep(3)
        
        for c in self.cookies_list:
            cookie = c.copy()
            cookie.pop('sameSite', None)
            cookie.pop('httpOnly', None) 
            cookie.pop('secure', None)
            cookie.setdefault('domain', '.facebook.com')
            try:
                self.driver.add_cookie(cookie)
            except: 
                pass
        
        self.driver.get("https://m.facebook.com")
        time.sleep(4)

    def load_post(self, post_url):
        print(f"Loading groups post: {post_url}")
        
        # For groups, try both mobile and mbasic
        urls_to_try = []
        
        if "groups/" in post_url:
            # Create both mobile and mbasic versions
            mobile_url = post_url.replace("mbasic.facebook.com", "m.facebook.com").replace("www.facebook.com", "m.facebook.com")
            mbasic_url = post_url.replace("m.facebook.com", "mbasic.facebook.com").replace("www.facebook.com", "mbasic.facebook.com")
            
            urls_to_try = [mobile_url, mbasic_url]
        else:
            urls_to_try = [post_url]
        
        for url_attempt in urls_to_try:
            try:
                print(f"Trying URL: {url_attempt}")
                self.driver.get(url_attempt)
                time.sleep(6)
                
                current_url = self.driver.current_url
                page_title = self.driver.title
                
                print(f"Current URL: {current_url}")
                print(f"Page title: {page_title}")
                
                # Detect layout
                if "m.facebook.com" in current_url:
                    self.current_layout = "mobile"
                elif "mbasic.facebook.com" in current_url:
                    self.current_layout = "mbasic"
                else:
                    self.current_layout = "www"
                
                print(f"Detected layout: {self.current_layout}")
                
                # Check login status
                if any(keyword in page_title.lower() for keyword in ["log in", "login", "đăng nhập"]):
                    print("❌ Not logged in with this URL, trying next...")
                    continue
                
                # Check if we can see group content
                group_indicators = self.driver.find_elements(By.XPATH, 
                    "//div[contains(text(),'group') or contains(text(),'nhóm') or contains(text(),'Group')]")
                
                if group_indicators or "groups/" in current_url:
                    print(f"✅ Successfully loaded groups post with {self.current_layout} layout")
                    return True
                else:
                    print("⚠️ No group indicators found, but proceeding...")
                    return True
                    
            except Exception as e:
                print(f"Failed to load {url_attempt}: {e}")
                continue
        
        print("❌ Failed to load post with any URL variant")
        return False

    def get_enhanced_user_info(self, element):
        """Enhanced method to extract user information from comment element"""
        user_info = {
            'name': 'Unknown',
            'uid': 'Unknown',
            'profile_url': '',
            'display_name': '',
            'verified': False
        }
        
        try:
            # Strategy 1: Look for profile links with enhanced selectors
            profile_selectors = [
                ".//a[contains(@href, 'profile.php')]",
                ".//a[contains(@href, 'user.php')]",
                ".//a[contains(@href, '/profile/')]",
                ".//strong/a[contains(@href, 'facebook.com/')]",
                ".//h3/a[contains(@href, 'facebook.com/')]",
                ".//span/a[contains(@href, 'facebook.com/')]",
                ".//div[@role='link']/a[contains(@href, 'facebook.com/')]",
                ".//a[contains(@data-sigil, 'profile')]",
                # Additional selectors for mobile layout
                ".//a[contains(@href, 'facebook.com/') and not(contains(@href, 'groups'))]",
                ".//strong[not(.//a)]//a[contains(@href, 'facebook.com/')]",
                ".//h3[not(.//a)]//a[contains(@href, 'facebook.com/')]",
                ".//span[not(.//a)]//a[contains(@href, 'facebook.com/')]",
                # More specific selectors for comment authors
                ".//div[contains(@class, 'comment')]//a[contains(@href, 'facebook.com/')]",
                ".//div[contains(@data-sigil, 'comment')]//a[contains(@href, 'facebook.com/')]",
                ".//article//a[contains(@href, 'facebook.com/')]",
                ".//div[@role='article']//a[contains(@href, 'facebook.com/')]"
            ]
            
            for selector in profile_selectors:
                try:
                    links = element.find_elements(By.XPATH, selector)
                    for link in links[:5]:  # Check first 5 links
                        link_text = link.text.strip()
                        link_href = link.get_attribute("href") or ""
                        
                        # Enhanced validation for username
                        if (link_text and 
                            2 <= len(link_text) <= 100 and 
                            not link_text.startswith('http') and
                            not link_text.isdigit() and
                            not any(ui in link_text.lower() for ui in ['like', 'reply', 'share', 'comment', 'thích', 'trả lời', 'chia sẻ', 'bình luận', 'groups', 'nhóm']) and
                            validate_username(link_text)):
                            
                            user_info['name'] = link_text
                            user_info['profile_url'] = link_href
                            
                            # Extract UID from URL
                            uid_info = extract_user_info_from_url(link_href)
                            if uid_info.get('profile_id'):
                                user_info['uid'] = uid_info['profile_id']
                            elif uid_info.get('user_id'):
                                user_info['uid'] = uid_info['user_id']
                            
                            # Check for verified badge
                            try:
                                verified_badge = link.find_elements(By.XPATH, ".//*[contains(@aria-label, 'verified') or contains(@title, 'verified')]")
                                if verified_badge:
                                    user_info['verified'] = True
                            except:
                                pass
                            
                            print(f"    DEBUG: Found user '{link_text}' with URL '{link_href}'")
                            break
                    
                    if user_info['name'] != 'Unknown':
                        break
                        
                except Exception as e:
                    print(f"Error with profile selector {selector}: {e}")
                    continue
            
            # Strategy 2: Look for display names in different elements
            if user_info['name'] == 'Unknown':
                display_selectors = [
                    ".//strong[not(.//a)]",
                    ".//h3[not(.//a)]",
                    ".//span[@data-sigil='comment-author']",
                    ".//div[@data-sigil='comment-author']",
                    ".//span[contains(@class, 'author')]",
                    ".//div[contains(@class, 'author')]",
                    ".//strong[contains(text(), ' ') and string-length(text()) > 3]",
                    ".//h3[contains(text(), ' ') and string-length(text()) > 3]",
                    # Additional selectors for comment authors
                    ".//div[contains(@class, 'comment')]//strong",
                    ".//div[contains(@data-sigil, 'comment')]//strong",
                    ".//article//strong",
                    ".//div[@role='article']//strong"
                ]
                
                for selector in display_selectors:
                    try:
                        elements = element.find_elements(By.XPATH, selector)
                        for elem in elements:
                            display_text = elem.text.strip()
                            if (display_text and 
                                2 <= len(display_text) <= 100 and
                                validate_username(display_text)):
                                user_info['display_name'] = display_text
                                if user_info['name'] == 'Unknown':
                                    user_info['name'] = display_text
                                print(f"    DEBUG: Found display name '{display_text}'")
                                break
                    except:
                        continue
            
            # Strategy 3: Extract from data attributes
            try:
                data_sigil = element.get_attribute("data-sigil") or ""
                if "comment" in data_sigil:
                    # Look for user info in data attributes
                    data_ft = element.get_attribute("data-ft") or ""
                    if data_ft:
                        # Try to parse JSON-like data
                        try:
                            import json
                            ft_data = json.loads(data_ft)
                            if 'actorID' in ft_data:
                                user_info['uid'] = str(ft_data['actorID'])
                                print(f"    DEBUG: Found UID from data-ft: {user_info['uid']}")
                        except:
                            pass
            except:
                pass
            
            # Strategy 4: Look for user mentions in text
            if user_info['name'] == 'Unknown':
                try:
                    full_text = element.text
                    # Look for patterns like "@username" or "Name:"
                    mention_patterns = [
                        r'@([a-zA-Z0-9_À-ỹ]+)',
                        r'^([A-Z][a-zÀ-ỹ]+(?:\s+[A-Z][a-zÀ-ỹ]+)*):',
                        r'Reply to ([A-Z][a-zÀ-ỹ]+(?:\s+[A-Z][a-zÀ-ỹ]+)*):',
                        r'([A-Z][a-zÀ-ỹ]+(?:\s+[A-Z][a-zÀ-ỹ]+)*)\s+[0-9]+'  # Name followed by numbers
                    ]
                    
                    for pattern in mention_patterns:
                        match = re.search(pattern, full_text)
                        if match:
                            potential_name = match.group(1)
                            if validate_username(potential_name):
                                user_info['name'] = potential_name
                                print(f"    DEBUG: Found name from mention pattern: '{potential_name}'")
                                break
                except:
                    pass
            
            # Cache user info for future use
            if user_info['uid'] != 'Unknown':
                self.user_cache[user_info['uid']] = user_info
            
        except Exception as e:
            print(f"Error extracting user info: {e}")
        
        return user_info

    def extract_enhanced_comment_content(self, element, username=""):
        """Enhanced comment content extraction with better text processing"""
        content_candidates = []
        
        try:
            # Get full text first
            full_text = element.text.strip()
            if not full_text or len(full_text) < 5:
                return ""
            
            print(f"    DEBUG: Full text: '{full_text[:100]}...'")
            
            # Strategy 1: Remove username and clean
            content = extract_comment_content(full_text, username)
            if content and len(content) >= 5:
                content_candidates.append(content)
                print(f"    DEBUG: Strategy 1 content: '{content[:50]}...'")
            
            # Strategy 2: Look for specific comment body elements
            comment_body_selectors = [
                ".//div[@data-sigil='comment-body']",
                ".//span[@data-sigil='comment-body']",
                ".//div[contains(@class, 'comment-body')]",
                ".//div[contains(@class, 'comment-text')]",
                ".//div[contains(@class, 'story_body')]",
                ".//div[contains(@data-ft, 'comment')]//div[not(.//a[contains(@href, 'profile.php')])]",
                # Additional selectors for mobile
                ".//div[contains(@class, 'comment')]//div[not(.//a)]",
                ".//div[contains(@data-sigil, 'comment')]//div[not(.//a)]",
                ".//article//div[not(.//a)]",
                ".//div[@role='article']//div[not(.//a)]"
            ]
            
            for selector in comment_body_selectors:
                try:
                    bodies = element.find_elements(By.XPATH, selector)
                    for body in bodies:
                        body_text = body.text.strip()
                        if body_text and body_text != username and len(body_text) >= 5:
                            cleaned_text = extract_comment_content(body_text, username)
                            if cleaned_text:
                                content_candidates.append(cleaned_text)
                                print(f"    DEBUG: Strategy 2 content: '{cleaned_text[:50]}...'")
                except:
                    continue
            
            # Strategy 3: Extract from text-only elements
            text_only_selectors = [
                ".//div[not(.//a) and not(.//button) and string-length(normalize-space(text())) > 20]",
                ".//span[not(.//a) and not(.//button) and string-length(normalize-space(text())) > 20]",
                ".//p[not(.//a) and not(.//button) and string-length(normalize-space(text())) > 20]",
                # Additional selectors for mobile
                ".//div[contains(@class, 'comment')]//div[not(.//a) and not(.//button)]",
                ".//div[contains(@data-sigil, 'comment')]//div[not(.//a) and not(.//button)]"
            ]
            
            for selector in text_only_selectors:
                try:
                    text_elements = element.find_elements(By.XPATH, selector)
                    for text_elem in text_elements:
                        text_content = text_elem.text.strip()
                        if text_content and text_content != username and len(text_content) >= 10:
                            cleaned_text = extract_comment_content(text_content, username)
                            if cleaned_text:
                                content_candidates.append(cleaned_text)
                                print(f"    DEBUG: Strategy 3 content: '{cleaned_text[:50]}...'")
                except:
                    continue
            
            # Strategy 4: Remove UI elements and extract content
            ui_patterns = [
                r'\b(Like|Reply|Share|Comment|Translate|Hide|Report|Block)\b',
                r'\b(Thích|Trả lời|Chia sẻ|Bình luận|Dịch|Ẩn|Báo cáo|Chặn)\b',
                r'\b\d+\s*(min|minutes?|hours?|days?|seconds?|phút|giờ|ngày|giây)\s*(ago|trước)?\b',
                r'\b(Top fan|Most relevant|Newest|All comments|Bình luận hàng đầu)\b',
                r'\b(View more|See more|Show more|Xem thêm|Hiển thị thêm)\b',
                r'\b(Write a comment|Viết bình luận)\b',
                r'^\d+\s*(like|love|reaction|thích|yêu|cảm xúc)\s*$',
                r'^(See translation|Xem bản dịch|Translate|Dịch)\s*$'
            ]
            
            # Apply UI cleaning to all candidates
            cleaned_candidates = []
            for candidate in content_candidates:
                cleaned = candidate
                for pattern in ui_patterns:
                    cleaned = re.sub(pattern, '', cleaned, flags=re.IGNORECASE)
                cleaned = re.sub(r'\s+', ' ', cleaned).strip()
                if cleaned and len(cleaned) >= 5:
                    cleaned_candidates.append(cleaned)
                    print(f"    DEBUG: Cleaned candidate: '{cleaned[:50]}...'")
            
            # Choose the best content
            if not cleaned_candidates:
                return ""
            
            # Prefer longer content, but not too long (likely contains UI elements)
            best_content = max(cleaned_candidates, key=lambda x: len(x) if len(x) < 500 else 0)
            
            # Final validation
            if self.is_ui_only_content(best_content):
                return ""
            
            print(f"    DEBUG: Final content: '{best_content[:100]}...'")
            return best_content.strip()
            
        except Exception as e:
            print(f"Error extracting comment content: {e}")
            return ""

    def is_ui_only_content(self, text):
        """Enhanced check for UI-only content - Less strict for mobile"""
        if not text or len(text.strip()) < 3:  # Reduced from 5 to 3
            return True
        
        text_clean = text.lower().strip()
        
        # UI patterns - Less strict for mobile layout
        ui_patterns = [
            r'^(like|reply|share|comment|translate|hide|report|block)(\s+\d+)?\s*$',
            r'^(thích|trả lời|chia sẻ|bình luận|dịch|ẩn|báo cáo|chặn)(\s+\d+)?\s*$',
            r'^\d+\s*(min|minutes?|hours?|days?|phút|giờ|ngày)\s*(ago|trước)?\s*$',
            r'^(top fan|most relevant|newest|all comments|view more|see more)\s*$',
            r'^(bình luận hàng đầu|xem thêm|hiển thị thêm)\s*$',
            r'^\d+\s*(like|love|reaction|thích|yêu|cảm xúc)\s*$',
            r'^(see translation|xem bản dịch|translate|dịch)\s*$',
            r'^(write a comment|viết bình luận|comment|bình luận)\s*$',
            r'^(group|nhóm|groups|các nhóm)\s*$',
            r'^[^\wÀ-ỹ]*$',  # Only punctuation/symbols
            r'^\d+$',  # Only numbers
            r'^[a-z]{1,2}\s*$'  # Very short text - reduced from 3 to 2
        ]
        
        for pattern in ui_patterns:
            if re.match(pattern, text_clean):
                return True
        
        # Check if it's just repeated characters - Less strict
        if len(set(text_clean)) <= 2 and len(text_clean) > 10:  # Increased threshold
            return True
        
        return False

    def determine_comment_type_enhanced(self, element, all_elements, index):
        """Enhanced comment type determination"""
        try:
            # Method 1: Check data attributes
            data_sigil = element.get_attribute("data-sigil") or ""
            if "reply" in data_sigil.lower():
                return "Reply"
            
            # Method 2: Check for reply indicators in text
            element_text = element.text.lower()
            reply_indicators = [
                'replied to', 'replying to', 'in reply to', 
                'trả lời', 'phản hồi', 'đáp lại',
                '@', 'reply:', 'phản hồi:', 'đáp:'
            ]
            if any(indicator in element_text for indicator in reply_indicators):
                return "Reply"
            
            # Method 3: Check indentation and positioning
            try:
                elem_location = element.location
                elem_size = element.size
                
                # Compare with nearby elements
                for compare_index in range(max(0, index-5), index):
                    if compare_index < len(all_elements):
                        try:
                            compare_elem = all_elements[compare_index]
                            compare_location = compare_elem.location
                            
                            # If this element is significantly more indented
                            if elem_location['x'] > compare_location['x'] + 20:
                                return "Reply"
                        except:
                            continue
            except:
                pass
            
            # Method 4: Check DOM structure
            try:
                # Look for reply-specific classes or attributes
                reply_classes = element.find_elements(By.XPATH, 
                    ".//*[contains(@class, 'reply') or contains(@class, 'comment-reply') or contains(@data-sigil, 'reply')]")
                if reply_classes:
                    return "Reply"
                
                # Check if nested within another comment
                comment_ancestors = element.find_elements(By.XPATH, 
                    "./ancestor::div[.//a[contains(@href,'profile.php')] and string-length(normalize-space(text())) > 50]")
                
                if len(comment_ancestors) > 0:
                    # Additional validation
                    for ancestor in comment_ancestors:
                        try:
                            ancestor_text = ancestor.text.strip()
                            element_text = element.text.strip()
                            
                            # If ancestor contains this element's text, it's likely a container
                            if element_text in ancestor_text and len(ancestor_text) > len(element_text) * 1.3:
                                return "Reply"
                        except:
                            continue
            except:
                pass
            
            # Default to main comment
            return "Comment"
            
        except Exception as e:
            print(f"Error determining comment type: {e}")
            return "Comment"

    def get_groups_selectors(self):
        """Get selectors optimized for Facebook Groups"""
        if self.current_layout == "mobile":
            return {
                'expand_links': [
                    # Mobile groups expand links
                    "//a[contains(text(),'View more comments')]",
                    "//a[contains(text(),'View previous comments')]",
                    "//a[contains(text(),'View more replies')]",
                    "//a[contains(text(),'Show more')]",
                    "//a[contains(text(),'See more')]",
                    "//a[contains(text(),'Xem thêm')]",
                    "//a[contains(text(),'Hiển thị thêm')]",
                    "//div[@role='button' and (contains(text(),'more') or contains(text(),'thêm'))]",
                    "//span[contains(text(),'View') and contains(text(),'more')]/ancestor::*[@role='button' or self::a][1]",
                    "//div[@data-sigil='more' or @data-sigil='expand']",
                    "//*[contains(@data-sigil,'comment')]//*[contains(text(),'more') or contains(text(),'thêm')]"
                ],
                'comment_containers': [
                    # Mobile groups comment containers
                    "//div[@data-sigil='comment']",
                    "//div[@data-sigil='comment-body']", 
                    "//div[contains(@data-ft, 'comment')]",
                    "//div[contains(@id, 'comment_')]",
                    "//article//div[.//a[contains(@href, 'profile.php')]]",
                    "//div[@role='article']//div[.//a[contains(@href, 'profile.php')]]",
                    "//div[contains(@class, 'story_body_container')]//div[.//a[contains(@href, 'profile.php')]]",
                    "//div[.//strong/a[contains(@href, 'profile.php')]]",
                    "//div[.//h3/a[contains(@href, 'profile.php')]]",
                    "//div[.//span/a[contains(@href, 'profile.php')]]",
                    # Broader selectors for groups
                    "//div[.//a[contains(@href, 'facebook.com/')] and string-length(normalize-space(text())) > 25]",
                    "//div[string-length(normalize-space(text())) > 40 and .//a[contains(@href, 'profile')]]"
                ]
            }
        else:
            # mbasic layout
            return {
                'expand_links': [
                    "//a[contains(text(),'View more comments')]",
                    "//a[contains(text(),'View previous comments')]",
                    "//a[contains(text(),'View more replies')]", 
                    "//a[contains(text(),'Show more')]",
                    "//a[contains(text(),'See more')]",
                    "//a[contains(text(),'Xem thêm')]",
                    "//a[contains(@href,'comment') and contains(text(),'more')]",
                    "//a[contains(@href,'reply') and contains(text(),'more')]"
                ],
                'comment_containers': [
                    "//div[@data-ft and contains(@data-ft, 'comment')]",
                    "//div[contains(@id, 'comment_')]",
                    "//table//div[.//a[contains(@href, 'profile.php')]]",
                    "//div[.//a[contains(@href, 'profile.php?id=')]]",
                    "//div[.//a[contains(@href, 'user.php?id=')]]",
                    "//div[.//h3/a[contains(@href, 'profile.php')]]"
                ]
            }

    def expand_groups_comments(self, max_iterations=120):
        """Specialized expansion for Facebook Groups"""
        print(f"=== EXPANDING GROUPS COMMENTS ({self.current_layout}) ===")
        
        selectors = self.get_groups_selectors()
        expand_selectors = selectors['expand_links']
        
        iteration = 0
        consecutive_failures = 0
        total_expansions = 0
        
        while iteration < max_iterations and consecutive_failures < 8:
            if self._stop_flag:
                break
                
            iteration += 1
            expanded_this_round = 0
            
            print(f"[Iteration {iteration}] Searching for expand links...")
            
            # Scroll to bottom
            self.driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
            time.sleep(random.uniform(2, 3))
            
            # Also try scrolling up a bit to trigger lazy loading
            if iteration % 3 == 0:
                self.driver.execute_script("window.scrollBy(0, -200);")
                time.sleep(1)
                self.driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
                time.sleep(2)
            
            # Find expandable elements
            expandable_elements = []
            for selector in expand_selectors:
                try:
                    elements = self.driver.find_elements(By.XPATH, selector)
                    for elem in elements:
                        try:
                            if elem.is_displayed() and elem.is_enabled():
                                elem_text = elem.text.strip().lower()
                                # Validate expand keywords
                                expand_keywords = ['view', 'more', 'show', 'see', 'xem', 'thêm', 'hiển thị', 'comment', 'reply', 'bình luận', 'phản hồi']
                                if any(keyword in elem_text for keyword in expand_keywords):
                                    if elem not in expandable_elements:
                                        expandable_elements.append(elem)
                        except:
                            continue
                except:
                    continue
            
            print(f"  Found {len(expandable_elements)} expandable elements")
            
            # Click expandable elements
            for i, elem in enumerate(expandable_elements):
                if self._stop_flag:
                    break
                    
                try:
                    elem_text = elem.text.strip()
                    print(f"    Attempting {i+1}/{len(expandable_elements)}: '{elem_text}'")
                    
                    # Scroll into view
                    self.driver.execute_script("arguments[0].scrollIntoView({behavior: 'smooth', block: 'center'});", elem)
                    time.sleep(1.5)
                    
                    # Try multiple click methods
                    click_success = False
                    
                    # Method 1: Regular click
                    try:
                        elem.click()
                        click_success = True
                    except:
                        pass
                    
                    # Method 2: JavaScript click
                    if not click_success:
                        try:
                            self.driver.execute_script("arguments[0].click();", elem)
                            click_success = True
                        except:
                            pass
                    
                    # Method 3: Action chains click
                    if not click_success:
                        try:
                            from selenium.webdriver.common.action_chains import ActionChains
                            ActionChains(self.driver).move_to_element(elem).click().perform()
                            click_success = True
                        except:
                            pass
                    
                    if click_success:
                        expanded_this_round += 1
                        print(f"    ✓ Successfully clicked")
                        time.sleep(random.uniform(2, 4))
                    else:
                        print(f"    ✗ All click methods failed")
                    
                except Exception as e:
                    print(f"    ✗ Error clicking: {e}")
                    continue
            
            total_expansions += expanded_this_round
            
            if expanded_this_round > 0:
                print(f"✓ Expanded {expanded_this_round} elements in iteration {iteration}")
                consecutive_failures = 0
            else:
                consecutive_failures += 1
                print(f"✗ No expansion in iteration {iteration} (consecutive failures: {consecutive_failures})")
        
        print(f"=== EXPANSION COMPLETE: {total_expansions} total expansions ===")
        
        # Final loading wait
        for _ in range(3):
            if self._stop_flag:
                break
            self.driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
            time.sleep(2)

    def extract_groups_comments_enhanced(self):
        """Enhanced comment extraction with improved user detection and content processing"""
        print(f"=== EXTRACTING ENHANCED GROUPS COMMENTS ({self.current_layout}) ===")
        
        # Save page for debugging
        try:
            with open(f"debug_groups_{self.current_layout}.html", "w", encoding="utf-8") as f:
                f.write(self.driver.page_source)
            print(f"Saved page to debug_groups_{self.current_layout}.html")
        except:
            pass
        
        selectors = self.get_groups_selectors()
        comment_selectors = selectors['comment_containers']
        
        # Collect all comment elements
        all_comment_elements = []
        
        for i, selector in enumerate(comment_selectors):
            try:
                elements = self.driver.find_elements(By.XPATH, selector)
                print(f"Groups selector {i+1}: Found {len(elements)} elements")
                
                for elem in elements:
                    if elem not in all_comment_elements:
                        all_comment_elements.append(elem)
                        
            except Exception as e:
                print(f"Groups selector {i+1} failed: {e}")
                continue
        
        print(f"Total unique comment elements: {len(all_comment_elements)}")
        
        # If still no elements, try very broad search
        if len(all_comment_elements) == 0:
            print("⚠️ No comments with standard selectors, trying emergency fallback...")
            
            emergency_selectors = [
                "//div[.//a[contains(@href, 'facebook.com/')] and string-length(normalize-space(text())) > 20]",
                "//div[string-length(normalize-space(text())) > 30]",
                "//*[contains(text(), 'Like') or contains(text(), 'Thích')]/ancestor::div[string-length(text()) > 40][1]",
                "//span[string-length(normalize-space(text())) > 25]/ancestor::div[1]",
                "//a[contains(@href,'profile.php')]/ancestor::div[string-length(text()) > 30][1]"
            ]
            
            for selector in emergency_selectors:
                try:
                    elements = self.driver.find_elements(By.XPATH, selector)
                    print(f"Emergency selector: Found {len(elements)} elements")
                    for elem in elements:
                        if elem not in all_comment_elements:
                            all_comment_elements.append(elem)
                    
                    # Stop if we found some elements
                    if len(all_comment_elements) > 10:
                        break
                except:
                    continue
        
        # Sort by position
        try:
            all_comment_elements.sort(key=lambda x: (x.location['y'], x.location['x']))
        except:
            pass
        
        comments = []
        seen_content = set()
        processed_elements = set()
        
        print(f"Processing {len(all_comment_elements)} potential comment elements...")
        
        # Process each element
        for i, element in enumerate(all_comment_elements):
            if self._stop_flag:
                break
                
            try:
                # Skip if already processed
                elem_id = id(element)
                if elem_id in processed_elements:
                    continue
                processed_elements.add(elem_id)
                
                print(f"\n--- Element {i+1}/{len(all_comment_elements)} ---")
                
                # Enhanced user info extraction
                user_info = self.get_enhanced_user_info(element)
                username = user_info['name']
                
                # Enhanced content extraction
                content = self.extract_enhanced_comment_content(element, username)
                
                if not content or len(content) < 5 or self.is_ui_only_content(content):
                    print(f"  ✗ Skipped: invalid content")
                    continue
                
                # Enhanced comment type determination
                comment_type = self.determine_comment_type_enhanced(element, all_comment_elements, i)
                
                # Try to extract username from content if still unknown
                if username == 'Unknown' and self.current_layout == "mobile":
                    extracted_name = self.extract_username_from_content(content)
                    if extracted_name:
                        username = extracted_name
                        user_info['name'] = extracted_name
                        # Update content to remove the name
                        content = self.remove_username_from_content(content, extracted_name)
                
                # Create comment data with enhanced information
                comment_data = {
                    "UID": user_info['uid'],
                    "Name": username,
                    "DisplayName": user_info['display_name'],
                    "Content": content,
                    "ProfileLink": user_info['profile_url'],
                    "Verified": user_info['verified'],
                    "Type": comment_type,
                    "Layout": self.current_layout,
                    "ContentLength": len(content),
                    "ElementIndex": i,
                    "RawText": element.text.strip()[:200]  # Add raw text for debugging
                }
                
                # Deduplication
                content_signature = f"{username}_{content[:40]}"
                if content_signature in seen_content:
                    print("  ✗ Skipped: duplicate")
                    continue
                seen_content.add(content_signature)
                
                comments.append(comment_data)
                print(f"  ✅ Added {comment_type}: {username} - {content[:50]}...")
                print(f"  📍 Profile: {user_info['profile_url']}")
                print(f"  🆔 UID: {user_info['uid']}")
                
            except Exception as e:
                print(f"  Error processing element {i}: {e}")
                continue
        
        # Final cleanup and organization
        final_comments = self.cleanup_groups_comments_enhanced(comments)
        
        print(f"\n=== ENHANCED GROUPS EXTRACTION COMPLETE ===")
        main_count = len([c for c in final_comments if c['Type'] == 'Comment'])
        reply_count = len([c for c in final_comments if c['Type'] == 'Reply'])
        print(f"Final results: {main_count} main comments + {reply_count} replies = {len(final_comments)} total")
        
        return final_comments

    def extract_username_from_content(self, content):
        """Extract username from content using various patterns"""
        if not content:
            return None
        
        # Patterns to extract username from content - Reordered for better matching
        patterns = [
            (r'^([A-Z][a-zA-ZÀ-ỹ\s]+)\s+(https?://)', 'Name with URL'),  # Must come before Name: content
            (r'^([A-Z][a-zA-ZÀ-ỹ\s]+):\s*(.+)', 'Name: content'),
            (r'^@([a-zA-Z0-9_À-ỹ]+)\s+(.+)', '@username content'),
            (r'^([A-Z][a-zA-ZÀ-ỹ\s]+)\s+[0-9]+\s+(.+)', 'Name 123 content'),
            (r'^([A-Z][a-zA-ZÀ-ỹ\s]+)\s+[0-9]+$', 'Name 123'),
            (r'^([A-Z][a-zA-ZÀ-ỹ\s]+)\s*$', 'Name only'),
            # Additional patterns for Vietnamese names without colon - More specific
            (r'^([A-Z][a-zA-ZÀ-ỹ\s]+)\s+(không|chính|bạn|em|cô|chú|anh|chị)\s+(.+)', 'Name with Vietnamese words'),
        ]
        
        for pattern, description in patterns:
            match = re.match(pattern, content)
            if match:
                potential_name = match.group(1).strip()
                if validate_username(potential_name):
                    return potential_name
        
        return None

    def remove_username_from_content(self, content, username):
        """Remove username from content"""
        if not content or not username:
            return content
        
        # Remove username from beginning
        if content.startswith(username):
            content = content[len(username):].strip()
        
        # Remove username with word boundaries
        content = re.sub(rf'\b{re.escape(username)}\b', '', content, count=1).strip()
        
        # Remove leading/trailing punctuation
        content = re.sub(r'^[:\s]+', '', content)
        content = re.sub(r'[\s:]+$', '', content)
        
        return content

    def cleanup_groups_comments_enhanced(self, comments):
        """Enhanced cleanup with better deduplication and validation"""
        print("=== CLEANING UP ENHANCED GROUPS COMMENTS ===")
        
        cleaned = []
        final_seen = set()
        
        for comment in comments:
            # Enhanced deduplication
            signatures = [
                f"{comment['Name']}_{comment['Content'][:30]}",
                comment['Content'][:60] if len(comment['Content']) > 60 else comment['Content'],
                f"{comment['UID']}_{comment['Content'][:25]}" if comment['UID'] != "Unknown" else None
            ]
            
            is_duplicate = any(sig in final_seen for sig in signatures if sig)
            
            # Enhanced validation - Less strict for mobile layout
            if (not is_duplicate and 
                comment['Content'] and 
                len(comment['Content']) >= 5 and  # Reduced from 8 to 5
                not self.is_ui_only_content(comment['Content'])):
                
                # For mobile layout, be more lenient with username validation
                if self.current_layout == "mobile" and comment['Name'] == 'Unknown':
                    # Try to extract name from content if possible
                    content = comment['Content']
                    # Look for patterns like "Name: content" or "@name content"
                    name_patterns = [
                        r'^([A-Z][a-zÀ-ỹ]+(?:\s+[A-Z][a-zÀ-ỹ]+)*):\s*(.+)',
                        r'^@([a-zA-Z0-9_À-ỹ]+)\s+(.+)',
                        r'^([A-Z][a-zÀ-ỹ]+(?:\s+[A-Z][a-zÀ-ỹ]+)*)\s+[0-9]+\s+(.+)'
                    ]
                    
                    for pattern in name_patterns:
                        match = re.match(pattern, content)
                        if match:
                            potential_name = match.group(1)
                            if validate_username(potential_name):
                                comment['Name'] = potential_name
                                comment['Content'] = match.group(2).strip()
                                break
                
                cleaned.append(comment)
                for sig in signatures:
                    if sig:
                        final_seen.add(sig)
        
        # Sort: main comments first, then replies
        cleaned.sort(key=lambda x: (
            x['Type'] == 'Reply',
            x.get('ElementIndex', 999)
        ))
        
        print(f"Final cleaned comments: {len(cleaned)}")
        return cleaned

    def is_ui_only_content(self, text):
        """Enhanced check for UI-only content - Less strict for mobile"""
        if not text or len(text.strip()) < 3:  # Reduced from 5 to 3
            return True
        
        text_clean = text.lower().strip()
        
        # UI patterns - Less strict for mobile layout
        ui_patterns = [
            r'^(like|reply|share|comment|translate|hide|report|block)(\s+\d+)?\s*$',
            r'^(thích|trả lời|chia sẻ|bình luận|dịch|ẩn|báo cáo|chặn)(\s+\d+)?\s*$',
            r'^\d+\s*(min|minutes?|hours?|days?|phút|giờ|ngày)\s*(ago|trước)?\s*$',
            r'^(top fan|most relevant|newest|all comments|view more|see more)\s*$',
            r'^(bình luận hàng đầu|xem thêm|hiển thị thêm)\s*$',
            r'^\d+\s*(like|love|reaction|thích|yêu|cảm xúc)\s*$',
            r'^(see translation|xem bản dịch|translate|dịch)\s*$',
            r'^(write a comment|viết bình luận|comment|bình luận)\s*$',
            r'^(group|nhóm|groups|các nhóm)\s*$',
            r'^[^\wÀ-ỹ]*$',  # Only punctuation/symbols
            r'^\d+$',  # Only numbers
            r'^[a-z]{1,2}\s*$'  # Very short text - reduced from 3 to 2
        ]
        
        for pattern in ui_patterns:
            if re.match(pattern, text_clean):
                return True
        
        # Check if it's just repeated characters - Less strict
        if len(set(text_clean)) <= 2 and len(text_clean) > 10:  # Increased threshold
            return True
        
        return False

    def scrape_all_comments_enhanced(self, limit=0, resolve_uid=True, progress_callback=None):
        """Enhanced main scraping orchestrator"""
        print(f"=== STARTING ENHANCED GROUPS SCRAPING ===")
        
        # Step 1: Expand all content
        self.expand_groups_comments()
        
        if self._stop_flag:
            return []
        
        # Step 2: Extract comments with enhanced methods
        comments = self.extract_groups_comments_enhanced()
        
        # Step 3: Apply limit
        if limit > 0:
            comments = comments[:limit]
        
        # Step 4: Progress reporting
        if progress_callback:
            progress_callback(len(comments))
        
        return comments

    def scrape_all_comments(self, limit=0, resolve_uid=True, progress_callback=None):
        """Main scraping orchestrator for groups - Fallback method"""
        print(f"=== STARTING GROUPS SCRAPING (FALLBACK) ===")
        
        # Step 1: Expand all content
        self.expand_groups_comments()
        
        if self._stop_flag:
            return []
        
        # Step 2: Extract comments with enhanced methods
        comments = self.extract_groups_comments_enhanced()
        
        # Step 3: Apply limit
        if limit > 0:
            comments = comments[:limit]
        
        # Step 4: Progress reporting
        if progress_callback:
            progress_callback(len(comments))
        
        return comments

    def close(self):
        try: 
            self.driver.quit()
        except: 
            pass

    def save_comments_to_excel_enhanced(self, comments, filename="facebook_groups_comments_enhanced.xlsx"):
        """Save comments to Excel with enhanced formatting"""
        if not comments:
            print("No comments to save")
            return
        
        try:
            import pandas as pd
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils.dataframe import dataframe_to_rows
            
            # Create DataFrame with the required structure
            data = []
            for comment in comments:
                row = {
                    'Cột 1': comment.get('Content', ''),
                    'Tên ACC': comment.get('Name', 'Unknown'),
                    'UID': comment.get('UID', 'Unknown'),
                    'UID COMMENT': comment.get('ProfileLink', ''),
                    'UID TỔNG': comment.get('UID', 'Unknown')  # Using UID as UID TỔNG
                }
                data.append(row)
            
            df = pd.DataFrame(data)
            
            # Create Excel file
            with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Comments', index=False)
                
                # Get the workbook and worksheet
                workbook = writer.book
                worksheet = writer.sheets['Comments']
                
                # Style the header row
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                header_alignment = Alignment(horizontal="center", vertical="center")
                
                for cell in worksheet[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                
                # Auto-adjust column widths
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    
                    adjusted_width = min(max_length + 2, 100)  # Max width 100
                    worksheet.column_dimensions[column_letter].width = adjusted_width
                
                # Add borders
                thin_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                
                for row in worksheet.iter_rows(min_row=1, max_row=len(comments)+1):
                    for cell in row:
                        cell.border = thin_border
                        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            
            print(f"✅ Comments saved to {filename}")
            print(f"📊 Total comments saved: {len(comments)}")
            
            # Print preview
            print("\n📋 PREVIEW OF SAVED DATA:")
            print("=" * 120)
            print(f"{'Cột 1':<50} {'Tên ACC':<20} {'UID':<20} {'UID COMMENT':<30}")
            print("=" * 120)
            
            for i, comment in enumerate(comments[:5]):  # Show first 5 comments
                content = comment.get('Content', '')[:45] + "..." if len(comment.get('Content', '')) > 45 else comment.get('Content', '')
                name = comment.get('Name', 'Unknown')
                uid = comment.get('UID', 'Unknown')
                profile = comment.get('ProfileLink', '')[:25] + "..." if len(comment.get('ProfileLink', '')) > 25 else comment.get('ProfileLink', '')
                
                print(f"{content:<50} {name:<20} {uid:<20} {profile:<30}")
            
            if len(comments) > 5:
                print(f"... and {len(comments) - 5} more comments")
            
            print("=" * 120)
            
        except ImportError as e:
            print(f"❌ Error: {e}")
            print("Please install required packages: pip install pandas openpyxl")
        except Exception as e:
            print(f"❌ Error saving to Excel: {e}")

    def print_comments_summary_enhanced(self, comments):
        """Print comments summary in the required format"""
        if not comments:
            print("No comments to display")
            return
        
        print("\n" + "=" * 120)
        print("📊 FACEBOOK GROUPS COMMENTS SUMMARY")
        print("=" * 120)
        print(f"{'Cột 1':<50} {'Tên ACC':<20} {'UID':<20} {'UID COMMENT':<30}")
        print("=" * 120)
        
        for i, comment in enumerate(comments, 1):
            content = comment.get('Content', '')
            name = comment.get('Name', 'Unknown')
            uid = comment.get('UID', 'Unknown')
            profile = comment.get('ProfileLink', '')
            
            # Truncate content if too long
            if len(content) > 45:
                content = content[:45] + "..."
            
            # Truncate profile if too long
            if len(profile) > 25:
                profile = profile[:25] + "..."
            
            print(f"{content:<50} {name:<20} {uid:<20} {profile:<30}")
            
            # Add separator every 10 comments
            if i % 10 == 0:
                print("-" * 120)
        
        print("=" * 120)
        print(f"📈 Total Comments: {len(comments)}")
        print(f"👥 Unique Users: {len(set(c.get('Name', '') for c in comments))}")
        print(f"🔗 Profiles Found: {len([c for c in comments if c.get('ProfileLink')])}")
        print("=" * 120)

# ----------------------------
# Enhanced Groups-Optimized GUI
# ----------------------------
class EnhancedFBGroupsAppGUI:
    def __init__(self, root):
        self.root = root
        root.title("🏘️ Enhanced FB Groups Comment Scraper")
        root.geometry("1100x950")
        root.configure(bg="#e8f5e8")

        # Main frame
        main_frame = tk.Frame(root, bg="#e8f5e8")
        main_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=15)

        # Header
        header_frame = tk.Frame(main_frame, bg="#e8f5e8")
        header_frame.pack(fill="x", pady=(0,20))
        
        title_label = tk.Label(header_frame, text="🏘️ Enhanced Facebook Groups Comment Scraper", 
                              font=("Arial", 20, "bold"), bg="#e8f5e8", fg="#2d5a2d")
        title_label.pack()
        
        subtitle_label = tk.Label(header_frame, text="✨ Cải tiến: Xác định tên người dùng chính xác hơn + Xử lý content tối ưu", 
                                 font=("Arial", 11), bg="#e8f5e8", fg="#5a5a5a")
        subtitle_label.pack(pady=(5,0))

        # Input section
        input_frame = tk.LabelFrame(main_frame, text="📝 Thông tin bài viết Groups", font=("Arial", 12, "bold"), 
                                   bg="#e8f5e8", fg="#2d5a2d", relief="groove", bd=2)
        input_frame.pack(fill="x", pady=(0,15))

        tk.Label(input_frame, text="🔗 Link bài viết trong Groups:", bg="#e8f5e8", font=("Arial", 10)).pack(anchor="w", padx=15, pady=(15,5))
        self.entry_url = tk.Entry(input_frame, width=100, font=("Arial", 9))
        self.entry_url.pack(fill="x", padx=15, pady=(0,10))

        tk.Label(input_frame, text="🍪 Cookie Facebook (để truy cập Groups):", bg="#e8f5e8", font=("Arial", 10)).pack(anchor="w", padx=15, pady=(5,5))
        self.txt_cookie = tk.Text(input_frame, height=4, font=("Arial", 8))
        self.txt_cookie.pack(fill="x", padx=15, pady=(0,15))

        # Enhanced Options section
        options_frame = tk.LabelFrame(main_frame, text="⚙️ Cấu hình nâng cao", font=("Arial", 12, "bold"), 
                                     bg="#e8f5e8", fg="#2d5a2d", relief="groove", bd=2)
        options_frame.pack(fill="x", pady=(0,15))
        
        opt_grid = tk.Frame(options_frame, bg="#e8f5e8")
        opt_grid.pack(fill="x", padx=15, pady=15)
        
        # Options grid
        tk.Label(opt_grid, text="📊 Số lượng comment:", bg="#e8f5e8").grid(row=0, column=0, sticky="w")
        self.entry_limit = tk.Entry(opt_grid, width=10)
        self.entry_limit.insert(0, "0")
        self.entry_limit.grid(row=0, column=1, sticky="w", padx=(10,20))
        tk.Label(opt_grid, text="(0 = tất cả)", bg="#e8f5e8", fg="#6c757d").grid(row=0, column=2, sticky="w")

        self.headless_var = tk.BooleanVar(value=False)  # Default to visible for groups debugging
        tk.Checkbutton(opt_grid, text="👻 Chạy ẩn", variable=self.headless_var, 
                      bg="#e8f5e8", font=("Arial", 9)).grid(row=1, column=0, sticky="w", pady=(10,0))

        self.resolve_uid_var = tk.BooleanVar(value=True)  # Default to True for enhanced features
        tk.Checkbutton(opt_grid, text="🆔 Lấy UID", variable=self.resolve_uid_var, 
                      bg="#e8f5e8", font=("Arial", 9)).grid(row=1, column=1, sticky="w", pady=(10,0))

        self.enhanced_extraction_var = tk.BooleanVar(value=True)
        tk.Checkbutton(opt_grid, text="🔍 Enhanced extraction", variable=self.enhanced_extraction_var, 
                      bg="#e8f5e8", font=("Arial", 9)).grid(row=1, column=2, sticky="w", pady=(10,0))

        # File section
        file_frame = tk.LabelFrame(main_frame, text="💾 Xuất kết quả", font=("Arial", 12, "bold"), 
                                  bg="#e8f5e8", fg="#2d5a2d", relief="groove", bd=2)
        file_frame.pack(fill="x", pady=(0,15))
        
        file_row = tk.Frame(file_frame, bg="#e8f5e8")
        file_row.pack(fill="x", padx=15, pady=15)
        
        self.entry_file = tk.Entry(file_row, width=70, font=("Arial", 9))
        self.entry_file.insert(0, "enhanced_facebook_groups_comments.xlsx")
        self.entry_file.pack(side="left", fill="x", expand=True)
        
        tk.Button(file_row, text="📁 Chọn", command=self.choose_file, 
                 bg="#17a2b8", fg="white", font=("Arial", 9)).pack(side="right", padx=(10,0))

        # Status section
        status_frame = tk.LabelFrame(main_frame, text="📊 Trạng thái thực thi", font=("Arial", 12, "bold"), 
                                    bg="#e8f5e8", fg="#2d5a2d", relief="groove", bd=2)
        status_frame.pack(fill="x", pady=(0,15))
        
        self.lbl_status = tk.Label(status_frame, text="✅ Sẵn sàng scrape Facebook Groups với tính năng nâng cao", fg="#28a745", 
                                  wraplength=1000, justify="left", font=("Arial", 11), bg="#e8f5e8")
        self.lbl_status.pack(anchor="w", padx=15, pady=(15,5))

        self.lbl_progress_detail = tk.Label(status_frame, text="💡 Cải tiến: Xác định tên người dùng chính xác hơn + Xử lý content tối ưu + Deduplication nâng cao", 
                                          fg="#6c757d", wraplength=1000, justify="left", font=("Arial", 9), bg="#e8f5e8")
        self.lbl_progress_detail.pack(anchor="w", padx=15, pady=(0,10))

        # Progress bar
        self.progress_bar = ttk.Progressbar(status_frame, mode='indeterminate')
        self.progress_bar.pack(fill="x", padx=15, pady=(0,15))

        # Control buttons
        button_frame = tk.Frame(main_frame, bg="#e8f5e8")
        button_frame.pack(fill="x", pady=20)
        
        self.btn_start = tk.Button(button_frame, text="🚀 Bắt đầu Enhanced Scraping", bg="#28a745", fg="white", 
                                  font=("Arial", 14, "bold"), command=self.start_scrape, 
                                  pady=12, padx=40)
        self.btn_start.pack(side="left")

        self.btn_stop = tk.Button(button_frame, text="⏹️ Dừng", bg="#dc3545", fg="white", 
                                 font=("Arial", 14, "bold"), command=self.stop_scrape, 
                                 state=tk.DISABLED, pady=12, padx=40)
        self.btn_stop.pack(side="left", padx=(25,0))

        self.progress_var = tk.IntVar(value=0)
        self.progress_label = tk.Label(button_frame, textvariable=self.progress_var, fg="#28a745", 
                                     font=("Arial", 18, "bold"), bg="#e8f5e8")
        self.progress_label.pack(side="right")

        self._scrape_thread = None
        self._stop_flag = False
        self.scraper = None

    def choose_file(self):
        f = filedialog.asksaveasfilename(
            defaultextension=".xlsx", 
            filetypes=[("Excel files", "*.xlsx"), ("CSV files", "*.csv")],
            title="Chọn file để lưu Enhanced Groups comments"
        )
        if f:
            self.entry_file.delete(0, tk.END)
            self.entry_file.insert(0, f)

    def start_scrape(self):
        url = self.entry_url.get().strip()
        cookie_str = self.txt_cookie.get("1.0", tk.END).strip()
        file_out = f"enhanced_groups_comments_{int(time.time())}.xlsx"
        
        if not url:
            messagebox.showerror("❌ Lỗi", "Vui lòng nhập link bài viết Groups!")
            return
        
        if not cookie_str:
            messagebox.showerror("❌ Lỗi", "Vui lòng nhập cookie Facebook!")
            return
        
        try: 
            limit = int(self.entry_limit.get().strip())
        except: 
            limit = 0

        self._stop_flag = False
        self.progress_var.set(0)
        self.progress_bar.start()
        self.lbl_status.config(text="🔄 Đang khởi động Enhanced Groups scraper...", fg="#fd7e14")
        self.lbl_progress_detail.config(text="⏳ Đang chuẩn bị trình duyệt với tính năng nâng cao...")
        self.btn_start.config(state=tk.DISABLED)
        self.btn_stop.config(state=tk.NORMAL)

        # Store parameters in instance variables for _scrape_worker to access
        self.scrape_url = url
        self.scrape_cookies = cookie_str
        self.scrape_file_out = file_out
        self.scrape_limit = limit
        self.scrape_headless = self.headless_var.get()
        self.scrape_resolve_uid = self.resolve_uid_var.get()
        self.scrape_enhanced = self.enhanced_extraction_var.get()

        self._scrape_thread = threading.Thread(target=self._scrape_worker)
        self._scrape_thread.daemon = True
        self._scrape_thread.start()

    def stop_scrape(self):
        self._stop_flag = True
        if hasattr(self, 'scraper') and self.scraper:
            self.scraper._stop_flag = True
        self.lbl_status.config(text="⏹️ Đang dừng Enhanced Groups scraper...", fg="#dc3545")
        self.btn_stop.config(state=tk.DISABLED)

    def _progress_cb(self, count):
        self.progress_var.set(count)
        self.lbl_status.config(text=f"📈 Đang xử lý Enhanced Groups... Đã lấy {count} comment/reply", fg="#28a745")
        self.root.update_idletasks()

    def _scrape_worker(self):
        """Worker thread for scraping"""
        try:
            self.lbl_status.config(text="🔄 Starting scraping...", fg="#fd7e14")
            self.progress_var.set(0)
            
            # Get parameters from instance variables
            url = self.scrape_url
            cookies = self.scrape_cookies
            limit = self.scrape_limit
            resolve_uid = self.scrape_resolve_uid
            enhanced = self.scrape_enhanced
            headless = self.scrape_headless
            
            if not url:
                self.lbl_status.config(text="❌ Please enter a URL", fg="#dc3545")
                return
            
            if not cookies:
                self.lbl_status.config(text="❌ Please enter cookies", fg="#dc3545")
                return
            
            # Initialize scraper
            self.lbl_status.config(text="🌐 Initializing scraper...", fg="#fd7e14")
            self.scraper = EnhancedFacebookGroupsScraper(cookies, headless=headless)
            
            if self._stop_flag:
                return
            
            # Load post
            self.lbl_status.config(text="📄 Loading post...", fg="#fd7e14")
            success = self.scraper.load_post(url)
            
            if not success:
                self.lbl_status.config(text="❌ Failed to load post", fg="#dc3545")
                return
            
            if self._stop_flag:
                return
            
            # Start scraping
            self.lbl_status.config(text="🔍 Scraping comments...", fg="#fd7e14")
            
            if enhanced:
                comments = self.scraper.scrape_all_comments_enhanced(limit=limit, resolve_uid=resolve_uid, progress_callback=self._progress_cb)
            else:
                comments = self.scraper.scrape_all_comments(limit=limit, resolve_uid=resolve_uid, progress_callback=self._progress_cb)
            
            if self._stop_flag:
                return
            
            if not comments:
                self.lbl_status.config(text="⚠️ No comments found", fg="#ffc107")
                return
            
            # Display results
            self.lbl_status.config(text=f"✅ Found {len(comments)} comments", fg="#28a745")
            
            # Print summary in required format
            self.scraper.print_comments_summary_enhanced(comments)
            
            # Save to Excel
            filename = f"facebook_groups_comments_{int(time.time())}.xlsx"
            self.scraper.save_comments_to_excel_enhanced(comments, filename)
            
            # Update GUI
            self.txt_result.delete("1.0", tk.END)
            self.txt_result.insert("1.0", f"✅ Scraping completed successfully!\n\n")
            self.txt_result.insert(tk.END, f"📊 Total comments: {len(comments)}\n")
            self.txt_result.insert(tk.END, f"👥 Unique users: {len(set(c.get('Name', '') for c in comments))}\n")
            self.txt_result.insert(tk.END, f"🔗 Profiles found: {len([c for c in comments if c.get('ProfileLink')])}\n")
            self.txt_result.insert(tk.END, f"💾 Saved to: {filename}\n\n")
            
            # Show sample data
            self.txt_result.insert(tk.END, "📋 SAMPLE DATA:\n")
            self.txt_result.insert(tk.END, "=" * 100 + "\n")
            self.txt_result.insert(tk.END, f"{'Cột 1':<40} {'Tên ACC':<15} {'UID':<15} {'UID COMMENT':<30}\n")
            self.txt_result.insert(tk.END, "=" * 100 + "\n")
            
            for comment in comments[:10]:  # Show first 10 comments
                content = comment.get('Content', '')[:35] + "..." if len(comment.get('Content', '')) > 35 else comment.get('Content', '')
                name = comment.get('Name', 'Unknown')
                uid = comment.get('UID', 'Unknown')
                profile = comment.get('ProfileLink', '')[:25] + "..." if len(comment.get('ProfileLink', '')) > 25 else comment.get('ProfileLink', '')
                
                self.txt_result.insert(tk.END, f"{content:<40} {name:<15} {uid:<15} {profile:<30}\n")
            
            if len(comments) > 10:
                self.txt_result.insert(tk.END, f"... and {len(comments) - 10} more comments\n")
            
            self.txt_result.insert(tk.END, "=" * 100 + "\n")
            
            self.progress_var.set(100)
            self.progress_bar.stop()
            
        except Exception as e:
            error_msg = f"❌ Error: {str(e)}"
            self.lbl_status.config(text=error_msg, fg="#dc3545")
            self.txt_result.delete("1.0", tk.END)
            self.txt_result.insert("1.0", error_msg)
            print(f"Scraping error: {e}")
        finally:
            if hasattr(self, 'scraper'):
                self.scraper.close()
            self.btn_start.config(state=tk.NORMAL)
            self.btn_stop.config(state=tk.DISABLED)
            self.progress_bar.stop()

# ----------------------------
# Run enhanced app
# ----------------------------
if __name__ == "__main__":
    root = tk.Tk()
    app = EnhancedFBGroupsAppGUI(root)
    root.mainloop()