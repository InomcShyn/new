#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Demo script để test output format theo yêu cầu
"""

import time

def create_sample_data():
    """Tạo dữ liệu mẫu theo format yêu cầu"""
    sample_comments = [
        {
            'Content': 'Bữa nay đo chiều cao cho con mà bất ngờ lun con cao thêm mấy phân các mom ơi #DinhDuongChatLuong #AbbottGrowGold',
            'Name': 'Lê Hải Minh',
            'UID': '100058456092012',
            'ProfileLink': 'https://www.facebook.com/groups/coconvuithayba/posts/1411362920261411/?comment_id=1414278199969883',
            'UID_TONG': '61560000765979'
        },
        {
            'Content': 'Cảm động quá và cũng thấy tự hào ❤️❤️❤️',
            'Name': 'Nguyễn Thị Mai',
            'UID': '100058456092013',
            'ProfileLink': 'https://www.facebook.com/nguyen.thi.mai.123',
            'UID_TONG': '61560000765980'
        },
        {
            'Content': 'Ông bố tiệm cơm 1k lại đánh bóng hình ảnh nữa à. Đ...',
            'Name': 'Trần Văn Bình',
            'UID': '100058456092014',
            'ProfileLink': 'https://www.facebook.com/tran.van.binh.456',
            'UID_TONG': '61560000765981'
        },
        {
            'Content': 'Mất tiền mua wifi đọc mấy bài như vầy mới đáng chớ...',
            'Name': 'Lê Thị Hương',
            'UID': '100058456092015',
            'ProfileLink': 'https://www.facebook.com/le.thi.huong.789',
            'UID_TONG': '61560000765982'
        },
        {
            'Content': 'Bát cháo này chất lượng quá 😍',
            'Name': 'Phạm Văn Cường',
            'UID': '100058456092016',
            'ProfileLink': 'https://www.facebook.com/pham.van.cuong.012',
            'UID_TONG': '61560000765983'
        }
    ]
    return sample_comments

def print_comments_summary(comments):
    """In summary theo format yêu cầu"""
    print("\n" + "=" * 120)
    print("📊 FACEBOOK GROUPS COMMENTS SUMMARY")
    print("=" * 120)
    print(f"{'Cột 1':<50} {'Tên ACC':<20} {'UID':<20} {'UID COMMENT':<30}")
    print("=" * 120)
    
    for i, comment in enumerate(comments, 1):
        content = comment.get('Content', '')
        name = comment.get('Name', 'Unknown')
        uid = comment.get('UID', 'Unknown')
        profile = comment.get('ProfileLink', '')
        
        # Truncate content if too long
        if len(content) > 45:
            content = content[:45] + "..."
        
        # Truncate profile if too long
        if len(profile) > 25:
            profile = profile[:25] + "..."
        
        print(f"{content:<50} {name:<20} {uid:<20} {profile:<30}")
        
        # Add separator every 5 comments
        if i % 5 == 0:
            print("-" * 120)
    
    print("=" * 120)
    print(f"📈 Total Comments: {len(comments)}")
    print(f"👥 Unique Users: {len(set(c.get('Name', '') for c in comments))}")
    print(f"🔗 Profiles Found: {len([c for c in comments if c.get('ProfileLink')])}")
    print("=" * 120)

def save_to_excel_format(comments, filename="demo_output.xlsx"):
    """Lưu vào Excel theo format yêu cầu"""
    try:
        import pandas as pd
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        # Create DataFrame với cấu trúc yêu cầu
        data = []
        for comment in comments:
            row = {
                'Cột 1': comment.get('Content', ''),
                'Tên ACC': comment.get('Name', 'Unknown'),
                'UID': comment.get('UID', 'Unknown'),
                'UID COMMENT': comment.get('ProfileLink', ''),
                'UID TỔNG': comment.get('UID_TONG', comment.get('UID', 'Unknown'))
            }
            data.append(row)
        
        df = pd.DataFrame(data)
        
        # Lưu vào Excel
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Comments', index=False)
            
            # Lấy workbook và worksheet
            workbook = writer.book
            worksheet = writer.sheets['Comments']
            
            # Style header row
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # Tự động điều chỉnh độ rộng cột
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 100)
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # Thêm borders
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            for row in worksheet.iter_rows(min_row=1, max_row=len(comments)+1):
                for cell in row:
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        
        print(f"✅ Comments saved to {filename}")
        print(f"📊 Total comments saved: {len(comments)}")
        
        # Print preview
        print("\n📋 PREVIEW OF SAVED DATA:")
        print("=" * 120)
        print(f"{'Cột 1':<50} {'Tên ACC':<20} {'UID':<20} {'UID COMMENT':<30}")
        print("=" * 120)
        
        for comment in comments[:5]:
            content = comment.get('Content', '')[:45] + "..." if len(comment.get('Content', '')) > 45 else comment.get('Content', '')
            name = comment.get('Name', 'Unknown')
            uid = comment.get('UID', 'Unknown')
            profile = comment.get('ProfileLink', '')[:25] + "..." if len(comment.get('ProfileLink', '')) > 25 else comment.get('ProfileLink', '')
            
            print(f"{content:<50} {name:<20} {uid:<20} {profile:<30}")
        
        print("=" * 120)
        
    except ImportError as e:
        print(f"❌ Error: {e}")
        print("Please install required packages: pip install pandas openpyxl")
    except Exception as e:
        print(f"❌ Error saving to Excel: {e}")

def main():
    """Chạy demo"""
    print("🧪 DEMO OUTPUT FORMAT")
    print("=" * 50)
    
    # Tạo dữ liệu mẫu
    comments = create_sample_data()
    
    # In summary
    print_comments_summary(comments)
    
    # Lưu vào Excel
    filename = f"demo_output_{int(time.time())}.xlsx"
    save_to_excel_format(comments, filename)
    
    print("\n🎉 Demo completed!")
    print("\n💡 Các tính năng đã được test:")
    print("  ✅ Output format theo yêu cầu")
    print("  ✅ Excel export với styling")
    print("  ✅ Console display đẹp")
    print("  ✅ Data structure chuẩn")

if __name__ == "__main__":
    main()